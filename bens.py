import conexao as cnx
import openpyxl as opnxl
from tqdm import tqdm

cur_origem = cnx.cur
cur_destino = cnx.cur_d
EMPRESA = 1

def moveis(nome_arquivo, nome_planilha):
    cur_destino.execute("update parampatri set EXER_ENC = null")
    cur_destino.execute("delete from pt_movbem")
    cur_destino.execute("delete from pt_cadpat")   
    cnx.commit()

    print('INSERINDO BENS...')

    planilha = opnxl.load_workbook(f"D:\Conversao\GIAP\DADOS\{nome_arquivo}")
    planilha_ativa = planilha[nome_planilha]

    insert = cur_destino.prep(
        """
        insert
        into
        pt_cadpat(
            codigo_pat
            , empresa_pat
            , codigo_gru_pat
            , chapa_pat
            , codigo_set_pat
            , codigo_set_atu_pat            
            , orig_pat
            , obs_pat
            , codigo_sit_pat
            , discr_pat
            , datae_pat
            , dtlan_pat
            , dt_contabil
            , valaqu_pat
            , valatu_pat
            , valres_pat
            , codigo_cpl_pat
            , codigo_tip_pat
            , codigo_for_pat
            , nota_pat  
            , responsa_pat
            , dae_pat
            , percentual_pat
            , percentemp_pat
            , percenqtd_pat
            )
        values(?, ?, ?, ?, ?, ?, ?,
               ?, ?, ?, ?, ?, ?, ?, 
               ?, ?, ?, ?, ?, ?, ?,
               ?, ?, ?, ?)""")

    for codigo_pat, row in enumerate(planilha_ativa.iter_rows(values_only=True)):
        if codigo_pat != 0:
            empresa_pat = EMPRESA
            codigo_gru_pat = row[5]
            chapa_pat = str(row[0]).zfill(6)
            codigo_set_pat = str(row[30]) + str(row[27])
            codigo_set_atu_pat = str(row[30]) + str(row[27])
            orig_pat =  row[68]
            obs_pat = row[42]
            codigo_sit_pat = row[35]
            discr_pat = row[42]
            datae_pat = row[1].strftime("%Y-%m-%d")
            dtlan_pat = row[1].strftime("%Y-%m-%d")
            dt_contabil = row[1].strftime("%Y-%m-%d")
            valaqu_pat = row[12]
            valatu_pat = row[75]
            try:
                valres_pat = round((valaqu_pat * row[13])/100, 2)
            except:
                valres_pat = None
            codigo_cpl_pat = row[6]
            try:
                codigo_tip_pat = cur_destino.execute(f"select codigo_tip from pt_cadtip where descricao_tip starting '{row[59][:40]}'").fetchone()[0]
            except:
                maxim = cur_destino.execute('select max(codigo_tip) + 1 from pt_cadtip').fetchone()[0]
                nome_tipo = row[59][:60]
                cur_destino.execute(f"insert into pt_cadtip(codigo_tip, empresa_tip, descricao_tip) values ({maxim},{EMPRESA},'{nome_tipo}')")
                codigo_tip_pat = maxim
            codigo_for_pat = row[47]
            nota_pat = row[58]
            responsa_pat = None
            dae_pat = 'V' if row[74] > 0  else 'N'
            try:
                percentual_pat = round(100 / row[22], 2)
            except:
                percentual_pat = None
            percentemp_pat = 'M'
            percenqtd_pat = row[22]
            cur_destino.execute(insert,(codigo_pat,empresa_pat,codigo_gru_pat,chapa_pat,codigo_set_pat,codigo_set_atu_pat,orig_pat,obs_pat,codigo_sit_pat,discr_pat[:255],
                                        datae_pat,dtlan_pat,dt_contabil,valaqu_pat,valatu_pat,valres_pat,codigo_cpl_pat,codigo_tip_pat,codigo_for_pat,nota_pat,responsa_pat,dae_pat,
                                        percentual_pat, percentemp_pat,percenqtd_pat))
    cnx.commit()

    cur_origem.execute("select id_patrimonio, cod_patrimonio from siop_sbs.pm_patrimonio pp")
    query = cur_origem.fetchall()
    total_rows = len(query)

    for row in tqdm(query, desc="INSERINDO ID ANTIGA:", total=total_rows):
        cur_destino.execute(f"UPDATE PT_CADPAT SET CODIGO_ANT_PAT = {row[0]} where chapa_pat = {str(row[1]).zfill(6)}")
    cnx.commit()

def aquisicao():
    cur_destino.execute("DELETE FROM PT_MOVBEM WHERE TIPO_MOV = 'A'")
    insert = cur_destino.prep(
            """
            insert
            into
            pt_movbem(
                codigo_mov
                , empresa_mov
                , codigo_pat_mov
                , data_mov
                , tipo_mov
                , valor_mov
                , codigo_cpl_mov
                , codigo_set_mov
                , documento_mov
                , dt_contabil
                , depreciacao_mov
                , codigo_bai_mov
                , historico_mov
            )
            values(?,?,?,?,?,?,?,?,?,?,?,?,?)
            """)
    
    codigo_mov = cur_destino.execute('select max(codigo_mov) from pt_movbem').fetchone()

    if codigo_mov == None:
        codigo_mov = codigo_mov.fetchone()
    else:
        codigo_mov = 0

    cur_destino.execute("SELECT CODIGO_PAT, DATAE_PAT, CODIGO_CPL_PAT, CODIGO_SET_PAT, VALAQU_PAT, NOTA_PAT FROM pt_cadpat")
    query = cur_destino.fetchall()
    total_rows = len(query)

    for row in tqdm(query, desc='PROGRESSO', total=total_rows):
        codigo_mov += 1
        empresa_mov = EMPRESA
        codigo_pat_mov = row[0]
        data_mov = row[1].strftime('%Y-%m-%d')
        tipo_mov = 'A'
        valor_mov = row[4]
        codigo_cpl_mov = row[2]
        codigo_set_mov = row[3]
        documento_mov = row[5]
        dt_contabil = data_mov
        depreciacao_mov = 'N'
        codigo_bai_mov = None
        historico_mov = 'AQUISIÇÃO DE PATRIMÔNIO'
        cur_destino.execute(insert,(codigo_mov,empresa_mov,codigo_pat_mov,data_mov,tipo_mov,valor_mov,codigo_cpl_mov,codigo_set_mov,documento_mov,dt_contabil,
                                    depreciacao_mov,codigo_bai_mov,historico_mov))
    cnx.commit()

def transferencias(nome_arquivo, nome_planilha):
    cur_destino.execute("DELETE FROM pt_movbem where tipo_mov = 'T'")

    try:
        cur_destino.execute('alter table pt_movbem add CODANT varchar(10)')
        cnx.commit()
    except:
        pass

    try:
        codigo_mov = cur_destino.execute("select max(codigo_mov) from pt_movbem").fetchone()[0]
    except:
        codigo_mov = 0    
    print("Inserindo Transferencias")

    planilha = opnxl.load_workbook(f"D:\Conversao\GIAP\DADOS\{nome_arquivo}")
    planilha_ativa = planilha[nome_planilha]
    total_rows = planilha_ativa.max_row - 1

    insert = cur_destino.prep(
            """
            insert
            into
            pt_movbem(
                codigo_mov
                , empresa_mov
                , codigo_pat_mov
                , data_mov
                , tipo_mov
                , valor_mov
                , lote_mov
                , codigo_cpl_mov
                , codigo_set_mov
                , historico_mov
                , dt_contabil
                , depreciacao_mov
                , codigo_bai_mov
                , codigo_com_mov
                , codant
            )
            values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """)

    for i, row in tqdm(enumerate(planilha_ativa.iter_rows(values_only=True)), desc='PROGRESSO', total=total_rows):
        if i != 0:
            codigo_mov += 1
            empresa_mov = EMPRESA
            codigo_pat_mov = cur_destino.execute(f"select codigo_pat from pt_cadpat where chapa_pat = {str(row[4]).zfill(6)}").fetchone()[0]
            data_mov = row[2].strftime("%Y-%m-%d")
            tipo_mov = 'T'
            valor_mov = 0
            lote_mov = None
            codigo_cpl_mov = row[17]
            codigo_set_mov = str(row[12]) + str(row[14])
            historico_mov = row[6]
            dt_contabil = row[2].strftime("%Y-%m-%d")
            depreciacao_mov = 'N'
            codigo_bai_mov = None
            codigo_com_mov = None
            codant = row[0]

            cur_destino.execute(insert,(codigo_mov, empresa_mov, codigo_pat_mov, data_mov, 
                                 tipo_mov, valor_mov, lote_mov, codigo_cpl_mov, 
                                 codigo_set_mov, historico_mov, dt_contabil,
                                 depreciacao_mov, codigo_bai_mov, codigo_com_mov, codant))
    cnx.commit()

def baixas(nome_arquivo, nome_planilha):
    cur_destino.execute("DELETE FROM PT_MOVBEM WHERE TIPO_MOV = 'B'")
    print('INSERINDO BAIXAS...')

    insert = cur_destino.prep(
            """
            insert
            into
            pt_movbem(
                codigo_mov
                , empresa_mov
                , codigo_pat_mov
                , data_mov
                , tipo_mov
                , valor_mov
                , lote_mov
                , codigo_cpl_mov
                , codigo_set_mov
                , documento_mov
                , dt_contabil
                , depreciacao_mov
                , codigo_bai_mov
                , historico_mov
                , codant
            )
            values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """)
    
    planilha = opnxl.load_workbook(f"D:\Conversao\GIAP\DADOS\{nome_arquivo}")
    planilha_ativa = planilha[nome_planilha]
    total_rows = planilha_ativa.max_row - 1

    codigo_mov = cur_destino.execute('select max(codigo_mov) from pt_movbem').fetchone()[0]

    for z, row in tqdm(enumerate(planilha_ativa.iter_rows(values_only=True)), desc='PROGRESSO', total=total_rows):
        if z != 0:
            codigo_mov += 1
            empresa_mov = EMPRESA
            codigo_pat_mov = cur_destino.execute(f"select codigo_pat from pt_cadpat where chapa_pat = {str(row[4]).zfill(6)}").fetchone()[0]
            data_mov = row[1].strftime('%Y-%m-%d')
            tipo_mov = 'B'
            valor_mov = -float(row[7])
            lote_mov = None
            codigo_cpl_mov = row[5]
            codigo_set_mov = cur_destino.execute(f"select codigo_set_atu_pat from pt_cadpat where codigo_pat = {codigo_pat_mov}").fetchone()[0]
            documento_mov = row[0]
            dt_contabil = row[1].strftime('%Y-%m-%d')
            depreciacao_mov = None
            codigo_bai_mov = row[3]
            historico_mov = row[8]
            codant = row[0]
            cur_destino.execute(insert,(codigo_mov,empresa_mov,codigo_pat_mov,data_mov,tipo_mov,valor_mov,lote_mov,codigo_cpl_mov,codigo_set_mov,documento_mov,dt_contabil,
                                        depreciacao_mov,codigo_bai_mov,historico_mov, codant))
            cur_destino.execute(f'''update pt_cadpat set dtpag_pat = '{data_mov}', codigo_bai_pat = {codigo_bai_mov} where codigo_pat = {codigo_pat_mov}''')
    cur_destino.execute("update pt_cadpat a set a.valatu_pat = (select sum(b.valor_mov) from pt_movbem b where a.codigo_pat = b.codigo_pat_mov)")
    cnx.commit()

def reavaliacao(nome_arquivo, nome_planilha):
    cur_destino.execute("delete from pt_movbem where tipo_mov = 'R' and depreciacao_mov = 'N'")
    cnx.commit()
    print('INSERINDO REAVALIACOES...')


    planilha = opnxl.load_workbook(f"D:\Conversao\GIAP\DADOS\{nome_arquivo}")
    planilha_ativa = planilha[nome_planilha]
    total_rows = planilha_ativa.max_row - 1

    codigo_mov = cur_destino.execute("select max(codigo_mov) from pt_movbem").fetchone()[0]

    insert = cur_destino.prep(
            """
            insert
            into
            pt_movbem(
                codigo_mov
                , empresa_mov
                , codigo_pat_mov
                , data_mov
                , tipo_mov
                , valor_mov
                , lote_mov
                , codigo_cpl_mov
                , codigo_set_mov
                , documento_mov
                , dt_contabil
                , depreciacao_mov
                , codigo_bai_mov
                , historico_mov
            )
            values(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """)
    
    for i, row in tqdm(enumerate(planilha_ativa.iter_rows(values_only=True)), desc='PROGRESSO:', total=total_rows):
        if i != 0:
            codigo_mov += 1
            empresa_mov = EMPRESA
            codigo_pat_mov = cur_destino.execute(f"select codigo_pat from pt_cadpat where chapa_pat = {str(row[0]).zfill(6)}").fetchone()[0]
            data_mov = row[6].strftime("%Y-%m-%d")
            tipo_mov = 'R'
            valor_mov = row[10]
            lote_mov = None
            codigo_cpl_mov = None
            codigo_set_mov = cur_destino.execute(f"select codigo_set_atu_pat from pt_cadpat where codigo_pat = {codigo_pat_mov}").fetchone()[0]
            documento_mov = None
            dt_contabil = data_mov
            depreciacao_mov = 'N'
            codigo_bai_mov = None
            historico_mov = row[4]
            cur_destino.execute(insert,(codigo_mov,empresa_mov,codigo_pat_mov,data_mov,tipo_mov,valor_mov,lote_mov,codigo_cpl_mov,codigo_set_mov,documento_mov,dt_contabil,
                                            depreciacao_mov,codigo_bai_mov,historico_mov))
    cnx.commit()

def depreciacoes():
    cur_destino.execute('''DELETE FROM pt_movbem WHERE tipo_mov = 'R' AND depreciacao_mov = 'S' ''')
    cnx.commit()
    print('INSERINDO DEPRECIAÇÕES')

    cur_origem.execute("""
        SELECT
            CAST(id_patrimonio AS INTEGER) AS codigo_pat,
            dat_sistema AS data_mov,
            vlr_nf AS valor_mov,
            des_especificacao AS descricao,
            id_unidade_local AS codigo_set,
            'A' AS tipo_mov,
            'N' AS depreciacao_mov
        FROM
            siop_sbs.pm_patrimonio
        WHERE
            num_instituicao = 1
        UNION ALL
        SELECT
            CAST(aval.id_patrimonio AS INTEGER) AS codigo_pat,
            aval.dat_sistema,
            COALESCE(aval.vlr_diferenca, 0),
            aval.his_avaliacao,
            pat.id_unidade_local,
            'R',
            'N'
        FROM
            siop_sbs.pm_avaliacao aval
        JOIN siop_sbs.pm_patrimonio pat ON
            pat.id_patrimonio = aval.id_patrimonio
            AND pat.num_instituicao = 1
        UNION ALL
        SELECT
            CAST(dep.id_patrimonio AS INTEGER) AS codigo_pat,
            dep.dat_depreciacao,
            dep.vlr_depreciado_round * -1,
            'Depreciação Mensal',
            pat.id_unidade_local,
            'R',
            'S'
        FROM
            siop_sbs.pm_depreciacao dep
        JOIN siop_sbs.pm_patrimonio pat ON
            pat.id_patrimonio = dep.id_patrimonio
            AND pat.num_instituicao = 1
        UNION ALL 
        SELECT
            CAST(transf.id_patrimonio AS INTEGER) AS codigo_pat,
            tt.dat_transferencia,
            0,
            tt.his_transferencia,
            transf.id_unidade_local_destino,
            'T',
            'N'
        FROM
            siop_sbs.PM_TRANSFERENCIA_ITEM transf
        JOIN siop_sbs.pm_transferencia tt ON
            tt.id_transferencia = transf.id_transferencia
        JOIN siop_sbs.pm_patrimonio pat ON
            pat.id_patrimonio = transf.id_patrimonio
            AND pat.num_instituicao = 1
        WHERE
            transf.id_unidade_local_origem <> transf.id_unidade_local_destino
    """)
    
    rows = cur_origem.fetchall()

    insert_data = []
    codigo_mov = cur_destino.execute('SELECT MAX(codigo_mov) FROM pt_movbem').fetchone()[0]
    
    for row in tqdm(rows, desc="PROGRESSO:", total=len(rows)):
        if row[6] == "S":
            codigo_mov += 1
            empresa_mov = EMPRESA
            codigo_pat_mov = cur_destino.execute(f"SELECT codigo_pat FROM pt_cadpat WHERE codigo_ant_pat = {row[0]}").fetchone()[0]
            data_mov = row[1].strftime("%Y-%m-%d")
            tipo_mov = 'R'
            valor_mov = float(row[2])
            lote_mov = None
            codigo_cpl_mov = None
            codigo_set_mov = cur_destino.execute(f"SELECT codigo_set_atu_pat FROM pt_cadpat WHERE codigo_pat = {codigo_pat_mov}").fetchone()[0]
            documento_mov = None
            dt_contabil = data_mov
            depreciacao_mov = 'S'
            codigo_bai_mov = None
            historico_mov = 'DEPRECIAÇÃO'

            insert_data.append((codigo_mov, empresa_mov, codigo_pat_mov, data_mov, tipo_mov, valor_mov,
                                lote_mov, codigo_cpl_mov, codigo_set_mov, documento_mov, dt_contabil,
                                depreciacao_mov, codigo_bai_mov, historico_mov))

    insert = cur_destino.prep(
        """
        INSERT INTO pt_movbem(
            codigo_mov,
            empresa_mov,
            codigo_pat_mov,
            data_mov,
            tipo_mov,
            valor_mov,
            lote_mov,
            codigo_cpl_mov,
            codigo_set_mov,
            documento_mov,
            dt_contabil,
            depreciacao_mov,
            codigo_bai_mov,
            historico_mov
        )
        VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """)

    cur_destino.executemany(insert, insert_data)
    cnx.commit()

            
